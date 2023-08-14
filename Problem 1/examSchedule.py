### Exam Schedule ###

import gurobipy as gp
from gurobipy import GRB
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill
import pandas as pd
import os

# Formulating model using gurobipy

################################################################### SETS AND INDICES ############################################################

# Number of student
p_S_cnt = 20

# Number of Courses
p_C_cnt = 10

# Number of days
p_D_cnt = 7

# Number of Sessions
p_Se_cnt = 2

# Number of Exam rooms
p_R_cnt = 4

# Set of all courses
s_Course = [c for c in range(p_C_cnt)]

# Set of all students
s_Student = [s for s in range(p_S_cnt)]

# Set of all days 
s_Day = [d for d in range(p_D_cnt)]

# Set of all sessions
s_Session = [se for se in range(p_Se_cnt)]

################################################################### PARAMETERS ###################################################################

# Reading course_student_records.xlsx file
df = pd.read_excel('course_student_records.xlsx')

# Student - course record
p_Record = df.iloc[0:10, 1:21] # Selecting the table
p_Record = p_Record.values.tolist() # changing dataframe into list

################################################################### START FORMULATION ############################################################

# Start the timer
start_time = datetime.now()

# Create a new Gurobi model
model = gp.Model()

################################################################### DECISION VARIABLES ########################################################### 

# Exam routine
v_Routine = model.addVars(p_C_cnt, p_D_cnt, p_Se_cnt, vtype = GRB.BINARY, name = "v_Routine")

# Penalty
v_penalty = model.addVars(p_S_cnt, p_D_cnt, vtype = GRB.BINARY, name = "v_penalty")

################################################################### OBJECTIVE ###################################################################

model.setObjective(gp.quicksum(v_penalty[s, d] for s in s_Student for d in s_Day), sense = GRB.MINIMIZE)

################################################################### CONSTRAINTS #################################################################

# Minimize the total number of penalty (which is number of time students are giving more than 1 exam in a day).
model.addConstrs(gp.quicksum(p_Record[c][s]*v_Routine[c, d, se] for c in s_Course) <= 1 for s in s_Student for d in s_Day for se in s_Session)

# The maximum number of examinations that can be scheduled at any one time is limited by the number of class rooms
model.addConstrs(gp.quicksum(v_Routine[c, d, se] for c in s_Course) <= p_R_cnt for d in s_Day for se in s_Session) 

# As far as possible a student needs to take only one examination per day
model.addConstrs(gp.quicksum(p_Record[c][s]*v_Routine[c, d, se] for se in s_Session for c in s_Course) <= 1 + v_penalty[s, d] for s in s_Student for d in s_Day)

# Exam for every course should be happened if the course is taken by a student
model.addConstrs(gp.quicksum(v_Routine[c, d, se] for d in s_Day for se in s_Session) >= p_Record[c][s] for c in s_Course for s in s_Student)

# Maximum time an exam of any course can happen is 1
model.addConstrs(gp.quicksum(v_Routine[c, d, se] for d in s_Day for se in s_Session) <= 1 for c in s_Course)

################################################################### SOLVE #######################################################################

model.optimize()
print('OPTIMAL = ', model.status == GRB.OPTIMAL)

# End the timer
end_time = datetime.now()

print("Started at - ", start_time, '\n')
print("\nEnded at - ", end_time)

################################################################### OUTPUT ###################################################################

if model.status == GRB.OPTIMAL:
    # print(model.objVal)
    # for d in s_Day:
    #     for c in s_Course:
    #         for se in s_Session:
    #             print(v_Routine[c, d, se].x, end = ' ')
    #         print(end = '\t')
    #     print()
    # Saving the output file
    print("Output saved on Exam_Schedule_Output.xlsx file")

    wb = openpyxl.Workbook()
    
    sheet = wb.active
    sheet.title = "Exam Schedule"
    sheet.cell(row = 1, column = 1).value = "Total Penalty"
    sheet.cell(row = 1, column = 2).value = model.objVal
    sheet.cell(row = 3, column = 1).value = "Solution Status"
    sheet.cell(row = 3, column = 2).value = "OPTIMAL"
    sheet.cell(row = 5, column = 1).value = "Run Time"
    sheet.cell(row = 5, column = 2).value = end_time-start_time
    
    sheet.cell(row = 8, column = 1).value = "Exam Schedule"
    sheet.cell(row = 9, column = 1).value = "Day"
    for d in s_Day:
        sheet.cell(row = 10+d, column = 1).value = f"Day_{d+1}"
    for se in s_Session:
        sheet.cell(row = 9, column = 2+se).value = f"Session_{se+1}"
    for d in s_Day:
        for se in s_Session:
            sheet.cell(row = 10+d, column = 2+se).value = "Course: "
            for c in s_Course:
                if v_Routine[c, d, se].x == 1:
                    sheet.cell(row = 10+d, column = 2+se).value += f"{c}, "
            
    wb.save("Exam_Schedule_Output.xlsx")
    os.startfile("Exam_Schedule_Output.xlsx")
    
else:
    print("Solver failed to find an optimal solution")
