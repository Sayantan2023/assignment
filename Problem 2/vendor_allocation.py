### Vendor Allocation ###

import gurobipy as gp
from gurobipy import GRB
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill
import pandas as pd
import os
import random
import numpy as np

# Formulating model using gurobipy

################################################################### SETS AND INDICES ############################################################

# Number of Items
p_ItemCnt = 15

# Number of Vendors
p_VendorCnt = 5

# Number of Slab
p_SlabCnt = 3

# Set of Items
s_Item = [i for i in range(p_ItemCnt)]

# Set of Vendors
s_Vendor = [s for s in range(p_VendorCnt)]

# Set of slabs
s_Slab = [d for d in range(p_SlabCnt)]

################################################################### PARAMETERS ###################################################################

# Reading the input file input.xlsx
df = pd.read_excel('input.xlsx')

# Price of unit item ‘i' for vendor ‘v’
p_UPrice = df.iloc[2:17, 8:13] # Selecting the table
p_UPrice = p_UPrice.values.tolist() # changing dataframe into list

# Quantity of item ‘i' provided by vendor ‘v’
p_Quantity = df.iloc[2:17, 1:6] # Selecting the table
p_Quantity = p_Quantity.values.tolist() # changing dataframe into list

# Demand of item ‘i'
p_Demand = df.iloc[2:17, 15] # Selecting the table
p_Demand = p_Demand.values.tolist() # changing dataframe into list

# Total bid price for vendor 'v'
BidPrice = [sum(p_UPrice[i][v]*p_Quantity[i][v] for i in range(p_ItemCnt)) for v in range(p_VendorCnt)]

# Total bid quantity for vendor 'v'
TotalQ = [sum(p_Quantity[i][v] for i in range(p_ItemCnt)) for v in range(p_VendorCnt)]

# Penalty cost for unit unsatisfied demand of item ‘i'
p_PenCost = [sum(p_UPrice[i][v] for v in range(p_VendorCnt)) for i in range(p_ItemCnt)]

# A vendor 'v' provides a discount of ‘x1[v]’and ‘x2[v]’ % on the total spend if the total spent
# allocated to the vendor is greater than or equal to ‘S1[v]’ and ‘S2[v]’ respectively
x1 = [10]*p_VendorCnt
x2 = [15]*p_VendorCnt
S1 = [1/3*sum(BidPrice)/p_VendorCnt]*p_VendorCnt
S2 = [2/3*sum(BidPrice)/p_VendorCnt]*p_VendorCnt

# A vendor 'v' provides a discount of ‘y1[v]’and ‘y2[v]’ % on the total spend if the total quantity
# allocated to the vendor is greater than or equal to ‘Q1[v]’ and ‘Q2[v]’ respectively
y1 = [5]*p_VendorCnt
y2 = [10]*p_VendorCnt
Q1 = [1/3*sum(TotalQ)/p_VendorCnt]*p_VendorCnt
Q2 = [2/3*sum(TotalQ)/p_VendorCnt]*p_VendorCnt

# M is sufficiently large positive number(Big M)
M = 10000000000

################################################################### START FORMULATION ############################################################

# Start the timer
start_time = datetime.now()

# Create a new Gurobi model
model = gp.Model()

################################################################### DECISION VARIABLES ########################################################### 

# Quantity of order for item ‘i' assigned to vendor ‘v’
v_OrderAssign = model.addVars(p_ItemCnt, p_VendorCnt, vtype = GRB.CONTINUOUS, name = "v_OrderAssign")

# Quantity of unsatisfied demand of item ‘i’
v_penalty = model.addVars(p_ItemCnt, vtype = GRB.CONTINUOUS, name = "v_penalty")

# Total spend allocated to vendor ‘v’
v_AlcSpent = model.addVars(p_VendorCnt, vtype = GRB.CONTINUOUS, name = "v_AlcSpent")

# Total quantity allocated to vendor ‘v’
v_AlcQunt = model.addVars(p_VendorCnt, vtype = GRB.CONTINUOUS, name = "v_AlcQunt")

# Total spend allocated to vendor ‘v’ after discount
v_DiscCost = model.addVars(p_VendorCnt, vtype = GRB.CONTINUOUS, name = "v_DiscCost")

# Spend discount availed in %  from vendor ‘v’
v_SpDisc = model.addVars(p_VendorCnt, vtype = GRB.CONTINUOUS, name = "v_SpDisc")

# Shipment discount availed in %  from vendor ‘v’
v_ShDisc = model.addVars(p_VendorCnt, vtype = GRB.CONTINUOUS, name = "v_ShDisc")

# Total discounted price on unit spend Price from vendor ‘v’.
v_TDisc = model.addVars(p_VendorCnt, vtype = GRB.CONTINUOUS, name = "v_ShDisc")

# Is equal to 1 if the Buyer is in slab ‘s’ for vendor ‘v’ for spent discount
# slab 0 : [0, S1); slab 1 : [S1, S2); slab 2 : [S2, ∞)
a = model.addVars(p_VendorCnt, p_SlabCnt, vtype = GRB.CONTINUOUS, name = "a")

# Is equal to 1 if the Buyer is in slab ‘s’ for vendor ‘v’ for shipment discount
# slab 0 : [0, Q1); slab 1 : [Q1, Q2); slab 2 : [Q2, ∞)
b = model.addVars(p_VendorCnt, p_SlabCnt, vtype = GRB.CONTINUOUS, name = "b")


################################################################### OBJECTIVE ###################################################################

model.setObjective(gp.quicksum(v_DiscCost[v] for v in s_Vendor) + gp.quicksum(v_penalty[i]*p_PenCost[i] for i in s_Item), sense = GRB.MINIMIZE)

################################################################### CONSTRAINTS #################################################################

# The order for an item to a vendor should not exceed vendor’s capacity for that item.
model.addConstrs(v_OrderAssign[i, v] <= p_Quantity[i][v] for i in s_Item for v in s_Vendor)

# The total allocation should not exceed the demand of that item. There is a possibility that some demand of items can remain unfulfilled so adding a penalty as soft constraint.
model.addConstrs(gp.quicksum(v_OrderAssign[i, v] for v in s_Vendor) + v_penalty[i] == p_Demand[i] for i in s_Item) 

# Total spend allocated to vendor ‘v’.
model.addConstrs(v_AlcSpent[v] == gp.quicksum(v_OrderAssign[i, v]*p_UPrice[i][v] for i in s_Item) for v in s_Vendor)

# Total quantity allocated to vendor ‘v’.
model.addConstrs(v_AlcQunt[v] == gp.quicksum(v_OrderAssign[i, v] for i in s_Item) for v in s_Vendor)

# Spend discount availed in %  from vendor ‘v’.
for v in s_Vendor:
    model.addConstr(gp.quicksum(a[v, s] for s in s_Slab) == 1)
    model.addConstr(v_AlcSpent[v] <= a[v, 0]*S1[v] + M*(1-a[v, 0]) - 0.001) # 0.001 is epsilon
    model.addConstr(v_AlcSpent[v] <= a[v, 1]*S2[v] + M*(1-a[v, 1]) - 0.001)
    model.addConstr(v_AlcSpent[v] >= a[v, 1]*S1[v] - M*(1-a[v, 1]))
    model.addConstr(v_AlcSpent[v] >= a[v, 2]*S2[v] - M*(1-a[v, 2]))
    model.addConstr(v_SpDisc[v] == 0 * a[v, 0] + x1[v] * a[v, 1] + x2[v] * a[v, 2])

# Shipment discount availed in %  from vendor ‘v’.
for v in s_Vendor:
    model.addConstr(gp.quicksum(b[v, s] for s in s_Slab) == 1)
    model.addConstr(v_AlcQunt[v] <= b[v, 0]*Q1[v] + M*(1-b[v, 0]) - 0.001)
    model.addConstr(v_AlcQunt[v] <= b[v, 1]*Q2[v] + M*(1-b[v, 1]) - 0.001)
    model.addConstr(v_AlcQunt[v] >= b[v, 1]*Q1[v] - M*(1-b[v, 1]))
    model.addConstr(v_AlcQunt[v] >= b[v, 2]*Q2[v] - M*(1-b[v, 2]))
    model.addConstr(v_ShDisc[v] == 0 * b[v, 0] + y1[v] * b[v, 1] + y2[v] * b[v, 2])

# Total discounted price on unit spend Price from vendor ‘v’.
model.addConstrs(v_TDisc[v] == (1-v_SpDisc[v]/100)*(1-v_ShDisc[v]/100) for v in s_Vendor)

# Total spend allocated to vendor ‘v’ after spent discount and shipment discount.
model.addConstrs(v_DiscCost[v] == v_AlcSpent[v] * v_TDisc[v] for v in s_Vendor)

################################################################### SOLVE #######################################################################

# Set the NonConvex parameter to 2 as it is an NLP
model.Params.NonConvex = 2

model.optimize()
print('OPTIMAL = ', model.status == GRB.OPTIMAL)

# End the timer
end_time = datetime.now()

print("Started at - ", start_time, '\n')
print("\nEnded at - ", end_time)

################################################################### OUTPUT ###################################################################

if model.status == GRB.OPTIMAL:
    
    print("Output saved on Vendor_Allocation_Output.xlsx file")

    wb = openpyxl.Workbook()
    
    sheet = wb.active
    sheet.title = "Vendor_Allocation"
    
    sheet.cell(row = 1, column = 1).value = "Item"
    for v in s_Vendor:
        sheet.cell(row = 1, column = 2+v).value = f"Vendor_{v+1}"
    for i in s_Item:
        sheet.cell(row = 2+i, column = 1).value = f"Item_{i+1}"
    for i in s_Item:
        for v in s_Vendor:
            sheet.cell(row = 2+i, column = 2+v).value = round(v_OrderAssign[i, v].x, 2)

    sheet.cell(row = 19, column = 1).value = "Unfulfilled Demand"
    for i in s_Item:
        sheet.cell(row = 20, column = 1+i).value = f"Item_{i+1}"
        sheet.cell(row = 21, column = 1+i).value = round(v_penalty[i].x, 2)
    
    sheet.cell(row = 25, column = 1).value = "Total cost"
    sheet.cell(row = 25, column = 2).value = round(sum(v_DiscCost[v].x for v in s_Vendor), 2)
            
    wb.save("Vendor_Allocation_Output.xlsx")
    os.startfile("Vendor_Allocation_Output.xlsx")

else:
    print("Solver failed to find an optimal solution")
